import os
import sys
import argparse
import re
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define a maximum timeout
TIMEOUT = 30

def log(msg):
    # Standard print is captured by GitHub Actions logs
    print(f"[LOG] {time.strftime('%Y-%m-%d %H:%M:%S')} - {msg}")

def start_driver():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--headless")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    
    # O Selenium 4.6+ já possui o Selenium Manager embutido, não precisamos mais do ChromeDriverManager
    driver = webdriver.Chrome(options=opts)
    return driver

def do_login(driver, wait, email, password):
    try:
        email_field = wait.until(EC.presence_of_element_located((By.NAME, "Email")))
        password_field = driver.find_element(By.NAME, "Senha")
        
        email_field.clear()
        email_field.send_keys(email)
        
        password_field.clear()
        password_field.send_keys(password)
        
        try:
            submit = driver.find_element(By.XPATH, "//*[self::button or self::input][contains(.,'Entrar') or contains(.,'Acessar')] | //button[@type='submit']")
            driver.execute_script("arguments[0].click();", submit)
        except:
            password_field.send_keys(Keys.ENTER)
        
        time.sleep(5)
        log("Login submetido.")
    except Exception as e:
        log(f"Falha ao realizar login: {e}")
        driver.quit()
        import sys
        sys.exit(1)

def gravar_planilha(data, filepath):
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelos Detalhados"
    
    headers = [
        "ID do Modelo", "Nome da Operação", "Emissor", "Remetente", "Destinatário",
        "Expedidor", "Recebedor", "Informar Expedidor", "Informar Recebedor",
        "Tipo Tomador", "Tomador Outros", "Estado de Origem", "Cidade de Origem",
        "Estado de Destino", "Cidade de Destino", "Finalidade", "CFOP",
        "Observação", "Ativa", "CST - ICMS", "% ICMS Fixa", "Redução Base Cálculo",
        "Somar ICMS no Frete", "CST IBS/CBS", "Classificação Tributária",
        "% IBS UF", "% CBS"
    ]
    ws.append(headers)
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx, item in enumerate(data, 2):
        model_id = item.get("ID", "")
        link_id = f"https://www.logcte.com.br/ModeloCte/Editar?id={model_id}" if model_id else ""
        row_data = [
            link_id, item.get("Nome", ""), item.get("Emissor", ""), item.get("Remetente", ""),
            item.get("Destinatario", ""), item.get("Expedidor", ""), item.get("Recebedor", ""),
            item.get("InformarExpedidor", ""), item.get("InformarRecebedor", ""),
            item.get("TipoTomador", ""), item.get("TomadorOutros", ""), item.get("UFOrigem", ""),
            item.get("CidadeOrigem", ""), item.get("UFDestino", ""), item.get("CidadeDestino", ""),
            item.get("Finalidade", ""), item.get("CFOP", ""), item.get("Observacao", ""),
            item.get("Ativo", ""), item.get("CST_ICMS", ""), item.get("AliquotaICMSFixa", ""),
            item.get("ReducaoBaseCalculo", ""), item.get("SomarICMSFrete", ""),
            item.get("CST_IBS_CBS", ""), item.get("ClassificacaoTributaria", ""),
            item.get("AliquotaIBSUF", ""), item.get("AliquotaCBS", "")
        ]
        ws.append(row_data)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40)

    wb.save(filepath)

def main():
    parser = argparse.ArgumentParser(description="Extração de Modelos LogCTe Cloud")
    parser.add_argument("--cnpj", required=True, help="CNPJ do emissor para buscar")
    args = parser.parse_args()

    email = os.environ.get("LOGCTE_EMAIL")
    password = os.environ.get("LOGCTE_PASSWORD")
    cnpj = args.cnpj

    if not email or not password:
        log("ERRO FATAL: As credenciais LOGCTE_EMAIL e LOGCTE_PASSWORD não estão configuradas nas variáveis de ambiente.")
        sys.exit(1)

    driver = start_driver()
    wait = WebDriverWait(driver, TIMEOUT)
    
    try:
        log("Acessando a página Home...")
        driver.get("https://www.logcte.com.br/Home/home")
        time.sleep(3)

        if "login" in driver.current_url.lower():
            log("Formulário de login detectado.")
            do_login(driver, wait, email, password)

        log("Navegando para a página de Modelos...")
        driver.get("https://www.logcte.com.br/ModeloCte")
        time.sleep(5)

        if "login" in driver.current_url.lower():
            do_login(driver, wait, email, password)

        try:
            btn_filtros = wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(@onclick, 'VerOcultarFiltros')] | //*[contains(text(), 'Filtros')]")))
            txt = btn_filtros.text.strip().lower()
            if "ver" in txt or "mostrar" in txt:
                driver.execute_script("arguments[0].click();", btn_filtros)
                time.sleep(2)
        except:
            pass

        log(f"Pesquisando o CNPJ: {cnpj}...")
        driver.execute_script("$('#Emissor').select2('open');")
        time.sleep(2)

        search_input = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input.select2-search__field")))
        search_input.clear()
        search_input.send_keys(cnpj)
        time.sleep(4)

        log("Pressionando ENTER para selecionar a transportadora sugerida...")
        search_input.send_keys(Keys.ENTER)
        time.sleep(2)

        option_text = ""
        try:
            option_text = driver.execute_script("return $('#Emissor').select2('data')[0].text;")
        except:
            pass
            
        clean_cnpj = re.sub(r'\D', '', cnpj)
        filename = f"Resultado_{clean_cnpj}.xlsx"
        log(f"Nome do arquivo final: {filename}")

        log("Clicando em Filtrar...")
        btn_filtrar = wait.until(EC.element_to_be_clickable((By.ID, "filterModeloCteBtn")))
        btn_filtrar.click()
        time.sleep(8)

        log("Alterando resultados por página para 1000...")
        select_length = wait.until(EC.presence_of_element_located((By.NAME, "def-tableModeloCte_length")))
        Select(select_length).select_by_value("1000")
        time.sleep(8)

        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
        if len(rows) == 1 and ("nenhum" in rows[0].text.lower() or "não encontrado" in rows[0].text.lower()):
            log("Nenhum modelo encontrado para o emissor informado.")
            sys.exit(0)

        model_ids = []
        for row in rows:
            try:
                edit_link = row.find_element(By.XPATH, ".//a[contains(@onclick, 'EditarModeloCte')]")
                onclick_val = edit_link.get_attribute("onclick")
                match = re.search(r"EditarModeloCte\((\d+)\)", onclick_val)
                if match:
                    model_ids.append(int(match.group(1)))
            except:
                pass

        total_models = len(model_ids)
        log(f"Total de {total_models} IDs mapeados. Iniciando detalhamento...")

        def get_input_val(id_name):
            try:
                val = driver.execute_script(f"var el = document.getElementById('{id_name}'); return el ? el.value : '';")
                return str(val).strip() if val else ""
            except:
                return ""
                
        def get_select_text(id_name):
            try:
                text = driver.execute_script(f"""
                    var el = document.getElementById('{id_name}');
                    if (!el) return '';
                    if (window.jQuery && jQuery(el).data('select2')) {{
                        var data = jQuery(el).select2('data');
                        if (data && data.length > 0) return data[0].text;
                    }}
                    if (el.options && el.selectedIndex >= 0) {{
                        return el.options[el.selectedIndex].text;
                    }}
                    return '';
                """)
                if not text:
                    sel = Select(driver.find_element(By.ID, id_name))
                    text = sel.first_selected_option.text
                    
                text = text.strip()
                if text.startswith("Default:"):
                    text = text[8:].strip()
                if "selecione" in text.lower():
                    return ""
                return text
            except:
                return ""
                
        def get_checkbox_val(id_name):
            try:
                return driver.execute_script(f"var el = document.getElementById('{id_name}'); return (el && el.checked) ? 'Sim' : 'Não';")
            except:
                return "Não"

        detailed_data = []
        for index, model_id in enumerate(model_ids):
            log(f"Acessando modelo {index+1}/{total_models} (ID: {model_id})...")
            driver.get(f"https://www.logcte.com.br/ModeloCte/Editar?id={model_id}")
            time.sleep(4)
            
            nome = get_input_val("Nome")
            if not nome:
                time.sleep(4)
                nome = get_input_val("Nome")
                
            model_info = {
                "ID": model_id,
                "Nome": nome,
                "Emissor": get_select_text("EmissorId"),
                "Remetente": get_select_text("RemetenteId"),
                "Destinatario": get_select_text("DestinatarioId"),
                "Expedidor": get_select_text("ExpedidorId"),
                "Recebedor": get_select_text("RecebedorId"),
                "InformarExpedidor": get_checkbox_val("InformarExpedidor"),
                "InformarRecebedor": get_checkbox_val("InformarRecebedor"),
                "TipoTomador": get_select_text("TipodoTomador"),
                "TomadorOutros": get_select_text("TomadorOutrosId"),
                "UFOrigem": get_select_text("UFOrigemId"),
                "CidadeOrigem": get_select_text("CidadeOrigemId"),
                "UFDestino": get_select_text("UFDestinoId"),
                "CidadeDestino": get_select_text("CidadeDestinoId"),
                "Finalidade": get_select_text("Finalidade"),
                "CFOP": get_select_text("NaturezaId"),
                "Observacao": get_input_val("Observacao"),
                "Ativo": get_checkbox_val("Ativa"),
                "CST_ICMS": get_select_text("TipoSituacaoTributariaICMS"),
                "AliquotaICMSFixa": get_input_val("AliquotaICMSFixa"),
                "ReducaoBaseCalculo": get_input_val("ReducaoBaseCalculo"),
                "SomarICMSFrete": get_checkbox_val("SomarValorICMSValorFrete"),
                "CST_IBS_CBS": get_select_text("CSTIBSCBS"),
                "ClassificacaoTributaria": get_select_text("ClassificacaoTributariaId"),
                "AliquotaIBSUF": get_input_val("AliquotaIBSUF"),
                "AliquotaCBS": get_input_val("AliquotaCBS")
            }
            detailed_data.append(model_info)
            # Salvar progressivamente
            gravar_planilha(detailed_data, filename)

        log("Execução concluída com sucesso!")

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
